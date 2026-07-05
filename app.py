import os
import sys
import time
import threading
from datetime import datetime
import customtkinter as ctk
import cv2
import numpy as np
from openpyxl import load_workbook, Workbook
from PIL import Image


class SynapseApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("SYNAPSE | Biometric Security Suite")
        self.geometry("1300x780")  # Expanded slightly to fit the table cleanly

        ctk.set_appearance_mode("Dark")
        self.configure(fg_color="#0A0F14")

        # --- State Variables ---
        self.is_camera_active = False
        self.camera_thread = None
        self.cap = None
        self.excel_file = "Attendance_Register.xlsx"
        self.known_faces_dir = "known_faces"
        self.widescreen_mode = True
        self.cooldown_tracker = {}  # Tracks logged timestamps to prevent duplicate records
        self.attendance_records = []  # Keeps track of entries for the UI table

        # Initialize OpenCV Haar Cascade Face Detector
        self.face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

        os.makedirs(self.known_faces_dir, exist_ok=True)
        self._ensure_excel_exists()

        # Layout grids configuration
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        # --- Sidebar Panel ---
        self.sidebar = ctk.CTkFrame(self, width=280, corner_radius=0, fg_color="#101720", border_color="#00F0FF",
                                    border_width=1)
        self.sidebar.grid(row=0, column=0, sticky="nsew")

        ctk.CTkLabel(self.sidebar, text="SYNAPSE", font=("Arial", 28, "bold"), text_color="#00F0FF").pack(pady=(35, 5))
        ctk.CTkLabel(self.sidebar, text="BIOMETRIC ENGINE", font=("Courier", 11, "bold"), text_color="#FF007F").pack(
            pady=(0, 20))

        self.ratio_switch = ctk.CTkSwitch(
            self.sidebar,
            text="Widescreen Mode (16:9)",
            font=("Courier", 11, "bold"),
            text_color="#00F0FF",
            progress_color="#FF007F",
            command=self.toggle_aspect_ratio
        )
        self.ratio_switch.select()
        self.ratio_switch.pack(pady=(0, 20))

        ctk.CTkButton(
            self.sidebar, text="Register Student", font=("Courier", 13, "bold"),
            fg_color="#142132", text_color="#00F0FF", border_color="#00F0FF", border_width=1,
            hover_color="#1A314B", height=45, command=self.register_ui
        ).pack(pady=10, padx=25, fill="x")

        ctk.CTkButton(
            self.sidebar, text="Start Attendance", font=("Courier", 13, "bold"),
            fg_color="#142132", text_color="#FF007F", border_color="#FF007F", border_width=1,
            hover_color="#2B1625", height=45, command=self.start_attendance
        ).pack(pady=10, padx=25, fill="x")

        ctk.CTkButton(
            self.sidebar, text="Stop Camera", font=("Courier", 13, "bold"),
            fg_color="#1D121F", text_color="#E0A96D", border_color="#E0A96D", border_width=1,
            hover_color="#2D1B30", height=45, command=self.stop_camera
        ).pack(pady=10, padx=25, fill="x")

        ctk.CTkButton(
            self.sidebar, text="See Excel Sheet", font=("Courier", 13, "bold"),
            fg_color="#142132", text_color="#00F0FF", border_color="#00F0FF", border_width=1,
            hover_color="#1A314B", height=45, command=self.open_excel
        ).pack(pady=10, padx=25, fill="x")

        ctk.CTkButton(
            self.sidebar, text="Exit Application", font=("Courier", 13, "bold"),
            fg_color="#2A0815", text_color="#FF007F", border_color="#FF007F", border_width=1,
            hover_color="#420C21", height=45, command=self.safe_exit
        ).pack(side="bottom", pady=35, padx=25, fill="x")

        # --- Main View Container Panel ---
        self.main_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.main_frame.grid(row=0, column=1, padx=25, pady=25, sticky="nsew")
        self.main_frame.grid_columnconfigure(0, weight=3)
        self.main_frame.grid_columnconfigure(1, weight=2)  # Space column for the data table
        self.main_frame.grid_rowconfigure(0, weight=3)
        self.main_frame.grid_rowconfigure(1, weight=1)

        # Left Column: Video Viewport
        self.video_display = ctk.CTkLabel(
            self.main_frame, text="SYSTEM OFFLINE\n\nAwait Optical Initialization Pipeline...",
            font=("Courier", 15, "bold"), text_color="#00F0FF",
            fg_color="#101720", corner_radius=12, border_color="#1A2635", border_width=2
        )
        self.video_display.grid(row=0, column=0, padx=(0, 10), sticky="nsew", pady=(0, 15))

        # Bottom Column: Console Log Box
        self.log_box = ctk.CTkTextbox(
            self.main_frame, font=("Courier", 13), fg_color="#101720", text_color="#00F0FF",
            border_color="#FF007F", border_width=1, corner_radius=12
        )
        self.log_box.grid(row=1, column=0, columnspan=2, sticky="nsew")

        # --- Right Column: Alternative Modern Data Table Panel ---
        self.table_frame = ctk.CTkScrollableFrame(
            self.main_frame, fg_color="#101720", border_color="#00F0FF", border_width=1,
            corner_radius=12, label_text="LIVE ATTENDANCE ROSTER", label_font=("Courier", 14, "bold"),
            label_text_color="#00F0FF"
        )
        self.table_frame.grid(row=0, column=1, padx=(10, 0), sticky="nsew", pady=(0, 15))

        # Draw Table Headers
        self._draw_table_headers()

        self.print_log("SYNAPSE Dashboard Live. Core Biometric Systems loaded.")

    def _ensure_excel_exists(self):
        if not os.path.exists(self.excel_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance"
            ws.append(["Timestamp", "Student ID", "Name", "Status"])
            wb.save(self.excel_file)

    def _draw_table_headers(self):
        # Clear frame completely first
        for widget in self.table_frame.winfo_children():
            widget.destroy()

        headers = ["TIME", "ID", "NAME", "STATUS"]
        for col_idx, text in enumerate(headers):
            lbl = ctk.CTkLabel(self.table_frame, text=text, font=("Courier", 12, "bold"), text_color="#FF007F")
            lbl.grid(row=0, column=col_idx, padx=12, pady=8, sticky="w")

    def print_log(self, text):
        t_stamp = datetime.now().strftime("%H:%M:%S")
        self.log_box.insert("0.0", f"[{t_stamp}] >> {text}\n")

    def toggle_aspect_ratio(self):
        self.widescreen_mode = self.ratio_switch.get() == 1
        mode_text = "Widescreen (16:9)" if self.widescreen_mode else "Standard (4:3)"
        self.print_log(f"CAMERA VIEWPORT CONFIG CHANGED: Rendering in {mode_text} format.")

    def log_attendance(self, student_id, name):
        now = time.time()

        # Explicitly check cooldown. If logged in last 60 seconds, reject insertion!
        if student_id in self.cooldown_tracker:
            if now - self.cooldown_tracker[student_id] < 60.0:
                return

        self.cooldown_tracker[student_id] = now
        timestamp_str = datetime.now().strftime("%H:%M:%S")

        # 1. Update Excel Sheet
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active
            full_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.append([full_timestamp, student_id, name, "Present"])
            wb.save(self.excel_file)
            self.print_log(f"LOG SUCCESS: {name} ({student_id}) marked present in Excel.")
        except PermissionError:
            self.print_log("DATALINK CRITICAL ERROR: Attendance file locked! Close Excel window immediately.")
        except Exception as e:
            self.print_log(f"DATALINK ERROR: {str(e)}")

        # 2. Update the clean In-App Live UI Table
        self.attendance_records.append({
            "time": timestamp_str,
            "id": student_id,
            "name": name,
            "status": "PRESENT"
        })
        self._refresh_ui_table()

    def _refresh_ui_table(self):
        # Redraw headers
        self._draw_table_headers()

        # Draw newest records at the top row layout positions
        for row_idx, record in enumerate(reversed(self.attendance_records), start=1):
            values = [record["time"], record["id"], record["name"], record["status"]]
            for col_idx, val in enumerate(values):
                color = "#00F0FF" if col_idx != 3 else "#00FF00"
                lbl = ctk.CTkLabel(self.table_frame, text=val, font=("Courier", 12), text_color=color)
                lbl.grid(row=row_idx, column=col_idx, padx=12, pady=4, sticky="w")

    def register_ui(self):
        reg = ctk.CTkToplevel(self)
        reg.title("Identity Provisioning Matrix")
        reg.geometry("350x290")
        reg.configure(fg_color="#101720")
        reg.attributes("-topmost", True)
        reg.resizable(False, False)

        ctk.CTkLabel(reg, text="BIOMETRIC REGISTRATION", font=("Courier", 14, "bold"), text_color="#00F0FF").pack(
            pady=20)
        name_in = ctk.CTkEntry(reg, placeholder_text="Full Student Name", width=240, fg_color="#0A0F14",
                               border_color="#FF007F", text_color="#FFFFFF")
        name_in.pack(pady=8)
        id_in = ctk.CTkEntry(reg, placeholder_text="Unique Card/Student ID", width=240, fg_color="#0A0F14",
                             border_color="#FF007F", text_color="#FFFFFF")
        id_in.pack(pady=8)

        ctk.CTkButton(
            reg, text="Capture Vector Images", font=("Courier", 12, "bold"),
            fg_color="#142132", text_color="#00F0FF", border_color="#00F0FF", border_width=1, hover_color="#1A314B",
            command=lambda: [self.capture_student(name_in.get(), id_in.get()), reg.destroy()]
        ).pack(pady=25)

    def capture_student(self, name, student_id):
        if not name or not student_id:
            self.print_log("ABORTION ERROR: Parameter entry validation failed.")
            return
        if self.is_camera_active:
            self.print_log("THREAD OVERLAP: Operational pipeline busy.")
            return

        self.is_camera_active = True
        self.camera_thread = threading.Thread(target=self._capture_logic, args=(name, student_id), daemon=True)
        self.camera_thread.start()

    def _capture_logic(self, name, student_id):
        clean_name = name.replace(" ", "_")
        path = os.path.join(self.known_faces_dir, f"{student_id}_{clean_name}")
        os.makedirs(path, exist_ok=True)

        if sys.platform == "win32":
            self.cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)
        else:
            self.cap = cv2.VideoCapture(0)

        start_t = time.time()
        while not self.cap.isOpened():
            if time.time() - start_t > 5.0:
                self.print_log("HARDWARE ERROR: Camera initialization failed.")
                self.after(0, self.stop_camera)
                return
            time.sleep(0.1)

        self.print_log(f"Initiating dynamic data acquisition loop for: {name}...")

        frame_count = 1
        while frame_count <= 25 and self.is_camera_active:
            ret, frame = self.cap.read()
            if not ret or frame is None:
                time.sleep(0.01)
                continue

            ui_frame = frame.copy()
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = self.face_cascade.detectMultiScale(gray, 1.2, 5, minSize=(80, 80))

            if len(faces) > 0:
                (x, y, w, h) = faces[0]
                face_crop = frame[y:y + h, x:x + w]
                cv2.imwrite(f"{path}/img_{frame_count}.jpg", face_crop)

                cv2.rectangle(ui_frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                cv2.putText(ui_frame, f"CAPTURING: {frame_count}/25", (x, y - 10), cv2.FONT_HERSHEY_DUPLEX, 0.6,
                            (0, 255, 0), 1)
                frame_count += 1
            else:
                cv2.putText(ui_frame, "POSITION FACE IN FRONT OF SENSOR", (30, 50), cv2.FONT_HERSHEY_DUPLEX, 0.7,
                            (0, 0, 255), 2)

            self.after(0, self._render_frame, ui_frame)
            time.sleep(0.08)

        self.after(0, self.stop_camera)
        self.print_log(f"PROVISION COMPLETE: Vector profiles compiled in: {path}")

    def start_attendance(self):
        if self.is_camera_active:
            self.print_log("THREAD OVERLAP: Operational pipeline busy.")
            return

        self.is_camera_active = True
        self.print_log("OPTICAL SEARCH ENGINE RUNNING: Initializing camera sensor hardware...")

        self.camera_thread = threading.Thread(target=self._attendance_thread_logic, daemon=True)
        self.camera_thread.start()

    def _attendance_thread_logic(self):
        if sys.platform == "win32":
            self.cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)
        else:
            self.cap = cv2.VideoCapture(0)

        start_t = time.time()
        while not self.cap.isOpened():
            if not self.is_camera_active:
                return
            if time.time() - start_t > 6.0:
                self.print_log("HARDWARE ERROR: Optical sensor failed to start.")
                self.after(0, self.stop_camera)
                return
            time.sleep(0.1)

        self.print_log("SENSOR ONLINE: Commencing fluid tracking sweeps.")

        while self.is_camera_active and self.cap is not None:
            ret, frame = self.cap.read()
            if not ret or frame is None:
                time.sleep(0.01)
                continue

            ui_frame = frame.copy()
            gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

            # Real-Time Face Detection coordinates sweep
            faces = self.face_cascade.detectMultiScale(gray_frame, scaleFactor=1.1, minNeighbors=5, minSize=(100, 100))

            if len(faces) == 0:
                self.after(0, self._render_frame, ui_frame)
                time.sleep(0.04)
                continue

            # Extract moving bounds coordinates of face
            (x, y, w, h) = faces[0]
            x1, y1, x2, y2 = x, y, x + w, y + h

            live_face_sample = frame[y1:y2, x1:x2]
            live_gray = cv2.cvtColor(live_face_sample, cv2.COLOR_BGR2GRAY)

            match_found = False
            matched_id = "UNKNOWN"
            matched_name = "UNIDENTIFIED"
            best_match_score = -1.0

            if os.path.exists(self.known_faces_dir) and live_gray.size > 0:
                for folder in os.listdir(self.known_faces_dir):
                    if "_" in folder:
                        s_id, s_name = folder.split("_", 1)
                        folder_path = os.path.join(self.known_faces_dir, folder)

                        for img_name in os.listdir(folder_path)[:4]:
                            img_path = os.path.join(folder_path, img_name)
                            saved_img = cv2.imread(img_path, cv2.IMREAD_GRAYSCALE)

                            if saved_img is not None:
                                try:
                                    saved_resized = cv2.resize(saved_img, (live_gray.shape[1], live_gray.shape[0]))
                                    res = cv2.matchTemplate(live_gray, saved_resized, cv2.TM_CCOEFF_NORMED)
                                    _, max_val, _, _ = cv2.minMaxLoc(res)

                                    if max_val > best_match_score:
                                        best_match_score = max_val
                                        if max_val > 0.50:
                                            matched_id = s_id
                                            matched_name = s_name.replace("_", " ")
                                            match_found = True
                                except Exception:
                                    pass

            color = (0, 255, 0) if match_found else (0, 0, 255)
            text = f"{matched_name.upper()} ({matched_id})" if match_found else "UNIDENTIFIED IDENTITY"

            # Draw tracking graphics boxes dynamically
            cv2.rectangle(ui_frame, (x1, y1), (x2, y2), color, 2)
            self._draw_hud_corners(ui_frame, x1, y1, x2, y2, color)
            cv2.putText(ui_frame, text, (x1, y1 - 15), cv2.FONT_HERSHEY_DUPLEX, 0.6, color, 2)

            if match_found:
                self.after(0, self.log_attendance, matched_id, matched_name)

            self.after(0, self._render_frame, ui_frame)
            time.sleep(0.04)

    def _render_frame(self, frame):
        try:
            if frame is None or not self.is_camera_active:
                return
            img_h, img_w, _ = frame.shape

            w = self.video_display.winfo_width()
            h = self.video_display.winfo_height()

            if w < 10 or h < 10:
                w, h = 800, 450 if self.widescreen_mode else 480

            frame_aspect = img_w / img_h
            display_aspect = w / h

            if display_aspect > frame_aspect:
                target_h = h
                target_w = int(h * frame_aspect)
            else:
                target_w = w
                target_h = int(w / frame_aspect)

            if target_w <= 0 or target_h <= 0:
                return

            cv2_img = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            pil_img = Image.fromarray(cv2_img)
            ctk_img = ctk.CTkImage(light_image=pil_img, dark_image=pil_img, size=(int(target_w), int(target_h)))

            self.video_display.configure(image=ctk_img, text="")
            self.video_display._image = ctk_img
        except Exception:
            pass

    def _draw_hud_corners(self, img, x1, y1, x2, y2, color, length=20, thickness=4):
        cv2.line(img, (x1, y1), (x1 + length, y1), color, thickness)
        cv2.line(img, (x1, y1), (x1, y1 + length), color, thickness)
        cv2.line(img, (x2, y1), (x2 - length, y1), color, thickness)
        cv2.line(img, (x2, y1), (x2, y1 + length), color, thickness)
        cv2.line(img, (x1, y2), (x1 + length, y2), color, thickness)
        cv2.line(img, (x1, y2), (x1, y2 - length), color, thickness)
        cv2.line(img, (x2, y2), (x2 - length, y2), color, thickness)
        cv2.line(img, (x2, y2), (x2, y2 - length), color, thickness)

    def stop_camera(self):
        self.is_camera_active = False
        if self.cap is not None:
            self.cap.release()
            self.cap = None
        try:
            self.video_display.configure(image="", text="SYSTEM OFFLINE\n\nAwait Optical Initialization Pipeline...")
            self.video_display._image = None
        except Exception:
            pass
        self.print_log("OPTICAL ENGINE DOWN: Frame loops terminated.")

    def open_excel(self):
        try:
            if sys.platform == "win32":
                os.startfile(self.excel_file)
            elif sys.platform == "darwin":
                os.system(f"open '{self.excel_file}'")
            else:
                os.system(f"xdg-open '{self.excel_file}'")
            self.print_log("DATALINK SUCCESS: Active directory workbook opened.")
        except Exception as e:
            self.print_log(f"DATALINK CRITICAL ERROR: System failed to launch file. {str(e)}")

    def safe_exit(self):
        self.is_camera_active = False
        if self.cap is not None:
            self.cap.release()
        try:
            self.destroy()
        except Exception:
            pass
        sys.exit(0)


if __name__ == "__main__":
    app = SynapseApp()
    app.mainloop()