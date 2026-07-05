import os
import json
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

REGISTRY_FILE = "student_registry.json"
EXCEL_FILE = "Attendance_Register.xlsx"


def load_student_registry():
    """Reads the JSON local student registry database."""
    if os.path.exists(REGISTRY_FILE):
        with open(REGISTRY_FILE, 'r') as f:
            try:
                return json.load(f)
            except Exception:
                return {}
    return {}


def save_student_to_registry(name, student_id):
    """Saves student records to the local JSON file database."""
    registry = load_student_registry()
    registry[name.strip().upper()] = student_id.strip().upper()
    with open(REGISTRY_FILE, 'w') as f:
        json.dump(registry, f, indent=4)
    print(f"[REGISTRY] Saved: {name.strip().upper()} -> {student_id.strip().upper()}")


def initialize_excel_spreadsheet():
    """Safely creates a fresh workbook if it doesn't exist, ensuring it is a valid zip archive."""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Master Registry"
        ws.append(["System Initialized", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        wb.save(EXCEL_FILE)
        wb.close()


def mark_attendance(identity_label):
    """
    Looks up student data, creates dynamic daily tabs in Excel,
    prevents duplicate daily entries, and applies corporate formatting styles.
    """
    initialize_excel_spreadsheet()

    identity_label = identity_label.strip().upper()
    student_registry = load_student_registry()
    student_id = student_registry.get(identity_label, "UNKNOWN_ID")

    now = datetime.now()
    current_date = now.strftime("%Y-%m-%d")
    current_time = now.strftime("%H:%M:%S")

    try:
        wb = load_workbook(EXCEL_FILE)
    except Exception:
        wb = Workbook()

    if current_date in wb.sheetnames:
        ws = wb[current_date]
    else:
        ws = wb.create_sheet(title=current_date)
        ws.append(["Student Name", "Student ID", "Date", "Time", "Status"])

    if "Master Registry" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Master Registry"]

    already_marked = False
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] == identity_label:
            already_marked = True
            break

    if already_marked:
        wb.close()
        return f"[INFO] Attendance already registered today for {identity_label}."

    ws.append([identity_label, student_id, current_date, current_time, "Present"])

    # --- DESIGN & VISUAL STYLING ENGINE ---
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F6AA5", end_color="1F6AA5", fill_type="solid")
    present_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
        row[1].alignment = center_align
        row[2].alignment = center_align
        row[3].alignment = center_align
        row[4].alignment = center_align
        row[4].fill = present_fill

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 5, 14)

    wb.save(EXCEL_FILE)
    wb.close()
    return f"[SUCCESS] Recorded attendance for {identity_label} at {current_time}."